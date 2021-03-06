# Library Management System
# Copyright (C) 2020 Andrey Shmaykhel, Alexander Solovyov, Timur Allayarov
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""
This module contains utility functions in main app.
"""

from openpyxl import Workbook

from django.utils.translation import gettext as _

from .models import Book, Lease


def build_xlsx():
    """
    Generates XLSX file.
    """
    workbook = Workbook()

    books_worksheet = workbook.active
    books_worksheet.title = _("Books")
    build_books_sheet(books_worksheet)

    lease_worksheet = workbook.create_sheet(_("Leases"))
    build_leases_sheet(lease_worksheet)

    return workbook


def build_books_sheet(worksheet):
    """
    Generates books data sheet.
    """
    worksheet.cell(row=1, column=1, value="ISBN")
    worksheet.cell(row=1, column=2, value=_("Name"))
    worksheet.cell(row=1, column=3, value=_("Authors"))
    worksheet.cell(row=1, column=4, value=_("Added date"))
    worksheet.cell(row=1, column=5, value=_("Count"))
    book_list = Book.objects.order_by('added_date')
    for row, book in enumerate(book_list, start=2):
        worksheet.cell(row=row, column=1, value=book.formatted_isbn())
        worksheet.cell(row=row, column=2, value=book.name)
        worksheet.cell(row=row, column=3, value=book.authors)
        worksheet.cell(row=row, column=4, value=book.added_date)
        worksheet.cell(row=row, column=5, value=book.count)
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['C'].width = 40
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 10


def build_leases_sheet(worksheet):
    """
    Generates leases data sheet.
    """
    worksheet.cell(row=1, column=1, value=_("ID"))
    worksheet.cell(row=1, column=2, value=_("Student"))
    worksheet.cell(row=1, column=3, value=_("Book ISBN"))
    worksheet.cell(row=1, column=4, value=_("Issue date"))
    worksheet.cell(row=1, column=5, value=_("Expire date"))
    worksheet.cell(row=1, column=6, value=_("Return date"))
    lease_list = Lease.objects.order_by('issue_date')
    for row, lease in enumerate(lease_list, start=2):
        worksheet.cell(row=row, column=1, value=str(lease.id))
        worksheet.cell(row=row, column=2, value=str(lease.student))
        worksheet.cell(row=row, column=3, value=lease.book.formatted_isbn())
        worksheet.cell(row=row, column=4, value=lease.issue_date)
        worksheet.cell(row=row, column=5, value=lease.expire_date)
        worksheet.cell(row=row, column=6, value=lease.return_date)
    worksheet.column_dimensions['A'].width = 40
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 20
