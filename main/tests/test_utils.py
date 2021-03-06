# Library Management System
# Copyright (C) 2020 Andrey Shmaykhel, Alexander Solovyov, Timur Allayarov
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

"""
This module contains tests of utility functions in main app.
"""

from openpyxl import Workbook

from django.test import TestCase
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.utils.translation import gettext as _

from main.utils import build_xlsx, build_books_sheet, build_leases_sheet
from main.models import Book

from .utils import isbn_list_6, student_credentials, create_student_lease


class BuildXlsxFuncTests(TestCase):
    """
    Tests checking build_xlsx() function.
    """

    def test_build_xlsx_returnes_valid_workbook(self):
        """
        build_xlsx() returnes workbook with 2 worksheets: Books and
        Leases.
        """
        workbook = build_xlsx()
        self.assertEqual(len(workbook.sheetnames), 2)
        self.assertEqual(workbook.sheetnames[0], _("Books"))
        self.assertEqual(workbook.sheetnames[1], _("Leases"))


class BuildBooksSheetFuncTests(TestCase):
    """
    Tests checking build_books_sheet() function.
    """

    def test_build_books_sheet_on_empty_db(self):
        """
        If DB is empty, books worksheet is empty.
        """
        workbook = Workbook()
        worksheet = workbook.active
        build_books_sheet(worksheet)
        self.assertEqual(worksheet['A1'].value, "ISBN")
        self.assertEqual(worksheet['B1'].value, _("Name"))
        self.assertEqual(worksheet['C1'].value, _("Authors"))
        self.assertEqual(worksheet['D1'].value, _("Added date"))
        self.assertEqual(worksheet['E1'].value, _("Count"))
        self.assertIsNone(worksheet['A2'].value)

    def test_build_books_sheet_on_nonempty_db(self):
        """
        If DB is not empty, books worksheet is not empty.
        """
        for count, isbn in enumerate(isbn_list_6, start=1):
            Book.objects.create(
                isbn=isbn, name=isbn, authors=isbn, count=count)

        workbook = Workbook()
        worksheet = workbook.active
        build_books_sheet(worksheet)
        self.assertEqual(worksheet['A1'].value, "ISBN")
        self.assertEqual(worksheet['B1'].value, _("Name"))
        self.assertEqual(worksheet['C1'].value, _("Authors"))
        self.assertEqual(worksheet['D1'].value, _("Added date"))
        self.assertEqual(worksheet['E1'].value, _("Count"))
        self.assertEqual(worksheet['A2'].value, "978-0-00-000000-2")
        self.assertEqual(worksheet['B3'].value, "9780000000019")
        self.assertEqual(worksheet['C4'].value, "9780000000026")
        self.assertGreater(
            worksheet['D5'].value,
            timezone.now() - timezone.timedelta(minutes=1))
        self.assertEqual(worksheet['E6'].value, 5)


class BuildLeasesSheetFuncTests(TestCase):
    """
    Tests checking build_leases_sheet() function.
    """

    def test_build_leases_sheet_on_empty_db(self):
        """
        If DB is empty, leases worksheet is empty.
        """
        workbook = Workbook()
        worksheet = workbook.active
        build_leases_sheet(worksheet)
        self.assertEqual(worksheet['A1'].value, _("ID"))
        self.assertEqual(worksheet['B1'].value, _("Student"))
        self.assertEqual(worksheet['C1'].value, _("Book ISBN"))
        self.assertEqual(worksheet['D1'].value, _("Issue date"))
        self.assertEqual(worksheet['E1'].value, _("Expire date"))
        self.assertEqual(worksheet['F1'].value, _("Return date"))
        self.assertIsNone(worksheet['A2'].value)

    def test_build_leases_sheet_on_nonempty_db(self):
        """
        If DB is not empty, leases worksheet is not empty.
        """
        student_user = get_user_model().objects.create_user(
            **student_credentials,
            first_name="Student",
            last_name="Smith")
        student_group = Group.objects.get_or_create(name="Student")[0]
        student_user.groups.add(student_group)

        for count, isbn in enumerate(isbn_list_6, start=1):
            Book.objects.create(
                isbn=isbn, name=isbn, count=count)
            create_student_lease(isbn)

        workbook = Workbook()
        worksheet = workbook.active
        build_leases_sheet(worksheet)
        self.assertEqual(worksheet['A1'].value, _("ID"))
        self.assertEqual(worksheet['B1'].value, _("Student"))
        self.assertEqual(worksheet['C1'].value, _("Book ISBN"))
        self.assertEqual(worksheet['D1'].value, _("Issue date"))
        self.assertEqual(worksheet['E1'].value, _("Expire date"))
        self.assertEqual(worksheet['F1'].value, _("Return date"))
        self.assertEqual(
            worksheet['A2'].value,
            str(Book.objects.get(pk='9780000000002')
                .lease_set.get(student=get_user_model()
                .objects.get_by_natural_key(
                    student_credentials['username'])).id))
        self.assertEqual(worksheet['B3'].value, "Student Smith [student1]")
        self.assertEqual(worksheet['C4'].value, "978-0-00-000002-6")
        self.assertGreater(
            worksheet['D5'].value,
            timezone.now() - timezone.timedelta(minutes=1))
        self.assertEqual(
            worksheet['E6'].value,
            (timezone.now() + timezone.timedelta(days=30)).date())
        self.assertIsNone(worksheet['F7'].value)
